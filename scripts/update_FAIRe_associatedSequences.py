#!/usr/bin/env python3
"""
Update FAIRe experimentRunMetadata associatedSequences from an SRA attributes table.

Builds associatedSequences URLs from SRA run / BioSample / BioProject accessions and
left-merges into FAIRe on filename (keeps all FAIRe rows).

Optional: ``--empty-file`` (e.g. ``empty_files_list.txt`` from ``find_empty_corrupted_files.py``):
one filename per line (basenames); optional tab/comma/space-separated second name; matching
rows get ``associatedSequences`` set to the literal string ``NA``.
"""

import argparse
import re
import os
import shutil
import sys

import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore


def read_sra_attributes(path: str) -> pd.DataFrame:
    """Load SRA attributes from TSV, CSV, or Excel."""
    if not os.path.isfile(path):
        raise FileNotFoundError(f"SRA attributes file not found: {path}")
    lower = path.lower()
    if lower.endswith(".xlsx"):
        return pd.read_excel(path, engine="openpyxl")
    if lower.endswith(".xls"):
        return pd.read_excel(path)
    if lower.endswith(".csv"):
        return pd.read_csv(path, encoding="utf-8")
    return pd.read_csv(path, sep="\t", encoding="utf-8")


def build_associated_sequences(row, col_acc: str, col_bs: str, col_bp: str) -> str:
    """Three NCBI URLs joined by ' | '."""

    def norm(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        s = str(v).strip()
        return s if s and s.lower() != "nan" else ""

    a, b, p = norm(row.get(col_acc)), norm(row.get(col_bs)), norm(row.get(col_bp))
    parts = []
    if a:
        parts.append(f"https://www.ncbi.nlm.nih.gov/sra/{a}")
    if b:
        parts.append(f"https://www.ncbi.nlm.nih.gov/biosample/{b}")
    if p:
        parts.append(f"https://www.ncbi.nlm.nih.gov/bioproject/{p}")
    return " | ".join(parts)


def normalize_filename(s) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    return str(s).strip()


def _add_empty_list_token(names: set, raw: str) -> None:
    """Add normalized name and, for paths, basename so FAIRe filename matches either form."""
    n = normalize_filename(raw)
    if not n:
        return
    names.add(n)
    if "/" in n or "\\" in n:
        base = normalize_filename(os.path.basename(n.replace("\\", "/")))
        if base:
            names.add(base)


def load_empty_filename_set(path: str) -> set:
    """
    Read a text file of filenames (same style as ``empty_files_list.txt`` from
    ``find_empty_corrupted_files.py``: one basename per line when that script finds
    empty/corrupted files).

    Blank lines are skipped. Text after ``#`` on a line is treated as a comment.
    Lines that are only a comment (e.g. ``# No empty/corrupted files found``) add
    no names.

    Each line may list one or two names separated by tab, comma, or whitespace.
    If a token looks like a path, both the full string and ``os.path.basename`` are
    registered for matching FAIRe ``filename``.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Empty-file list not found: {path}")
    names = set()
    with open(path, encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.split("#", 1)[0].strip()
            if not line:
                continue
            if "\t" in line:
                parts = [p.strip() for p in line.split("\t")]
            elif "," in line:
                parts = [p.strip() for p in line.split(",")]
            else:
                parts = [p.strip() for p in re.split(r"\s+", line) if p.strip()]
            for p in parts:
                _add_empty_list_token(names, p)
    return names


def coalesce_sra_into_faire(sra_series: pd.Series, faire_series: pd.Series) -> pd.Series:
    out = faire_series.copy()
    nonempty = (
        sra_series.notna()
        & (sra_series.astype(str).str.strip() != "")
        & (sra_series.astype(str).str.lower() != "nan")
    )
    out.loc[nonempty] = sra_series.loc[nonempty]
    return out


def write_experiment_run_metadata_sheet(path: str, merged: pd.DataFrame) -> None:
    """Replace experimentRunMetadata body from row 3 (header) downward; keep rows 1–2."""
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")

    wb = load_workbook(path, read_only=False, data_only=False)
    if "experimentRunMetadata" not in wb.sheetnames:
        wb.close()
        raise ValueError("Workbook has no sheet 'experimentRunMetadata'")
    ws = wb["experimentRunMetadata"]

    # Remove old content from row 3 through end (keep preamble rows 1–2)
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)

    # Header row = Excel row 3
    for col_idx, name in enumerate(merged.columns, start=1):
        ws.cell(row=3, column=col_idx, value=name)

    # Data from row 4
    for r, row in enumerate(merged.itertuples(index=False), start=4):
        for c, val in enumerate(row, start=1):
            v = val
            if pd.isna(v):
                v = None
            elif isinstance(v, pd.Timestamp):
                v = v.to_pydatetime()
            ws.cell(row=r, column=c, value=v)

    wb.save(path)
    wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Merge SRA-derived associatedSequences into FAIRe experimentRunMetadata (on filename)."
    )
    parser.add_argument(
        "--FAIReMetadata",
        required=True,
        help="Path to FAIRe metadata Excel (.xlsx)",
    )
    parser.add_argument(
        "--SRA-attributes",
        dest="sra_attributes",
        required=True,
        help="Path to SRA attributes table (TSV/CSV/XLSX)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Path to write updated FAIRe Excel (default: overwrite --FAIReMetadata)",
    )
    parser.add_argument(
        "--empty-file",
        dest="empty_file",
        default=None,
        help=(
            "Optional path to a list file (e.g. empty_files_list.txt from "
            "find_empty_corrupted_files.py): one basename per line, optional pair "
            "on a line; # comments allowed. Matching FAIRe filename -> associatedSequences=NA."
        ),
    )
    args = parser.parse_args()
    faire_path = args.FAIReMetadata
    out_path = args.output or faire_path

    if not os.path.isfile(faire_path):
        print(f"Error: FAIRe metadata not found: {faire_path}", file=sys.stderr)
        return 1

    if not str(faire_path).lower().endswith(".xlsx"):
        print("Error: --FAIReMetadata must be an .xlsx file.", file=sys.stderr)
        return 1

    sra_df = read_sra_attributes(args.sra_attributes)

    col_acc, col_bs, col_bp = "accession", "biosample_accession", "bioproject_accession"
    need = [col_acc, col_bs, col_bp, "filename"]
    missing = [c for c in need if c not in sra_df.columns]
    if missing:
        print(
            f"Error: SRA attributes file must contain columns: {need}. Missing: {missing}",
            file=sys.stderr,
        )
        return 1

    sra_df = sra_df.copy()
    sra_df["associatedSequences"] = sra_df.apply(
        lambda r: build_associated_sequences(r, col_acc, col_bs, col_bp), axis=1
    )
    sra_merge = sra_df[["filename", "associatedSequences"]].copy()
    sra_merge["_sra_fn"] = sra_merge["filename"].map(normalize_filename)

    erm = pd.read_excel(
        faire_path,
        sheet_name="experimentRunMetadata",
        header=2,
        engine="openpyxl",
        keep_default_na=False,
    )
    if "filename" not in erm.columns:
        print(
            "Error: experimentRunMetadata must contain column 'filename'.",
            file=sys.stderr,
        )
        return 1
    if "associatedSequences" not in erm.columns:
        erm["associatedSequences"] = ""

    erm["_fn"] = erm["filename"].map(normalize_filename)
    merged = erm.merge(
        sra_merge.drop(columns=["filename"]).rename(
            columns={"associatedSequences": "associatedSequences_sra"}
        ),
        left_on="_fn",
        right_on="_sra_fn",
        how="left",
    )
    merged["associatedSequences"] = coalesce_sra_into_faire(
        merged["associatedSequences_sra"], merged["associatedSequences"]
    )
    merged = merged.drop(
        columns=[c for c in ("_fn", "_sra_fn", "associatedSequences_sra") if c in merged.columns]
    )

    if args.empty_file:
        try:
            empty_names = load_empty_filename_set(args.empty_file)
        except FileNotFoundError as e:
            print(f"Error: {e}", file=sys.stderr)
            return 1
        if empty_names:
            fn_norm = merged["filename"].map(normalize_filename)
            mask = fn_norm.isin(empty_names)
            merged.loc[mask, "associatedSequences"] = "NA"

    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    if os.path.abspath(out_path) != os.path.abspath(faire_path):
        shutil.copy2(faire_path, out_path)

    target = out_path
    write_experiment_run_metadata_sheet(target, merged)

    print(f"Updated sheet 'experimentRunMetadata' in: {target}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
